# Using openpyxl library to read from the file
import openpyxl


def print_active_worksheet(worksheet):
    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if i == 0:
                string = "" if col[i].value == "None" else col[i].value
                print(string, end="\t")
            else:
                print(col[i].value, end="\t")
        print('')


def generate_list_of_jobs(num_WS1, num_WS2, num_WS3, num_job):
    # workbook = openpyxl.load_workbook("../data/DATA.xlsx")
    file = input(
        "Masukkan nama file (pastikan file .xlsx sudah berada di folder data) (tanpa .xlsx) : ")
    workbook = openpyxl.load_workbook(f"../data/{file}.xlsx")

    # Define variable to read the active sheet:
    ws = workbook.active

    # Number of workspaces in a WS and number of jobs
    n_WS1 = num_WS1
    n_WS2 = num_WS2
    n_WS3 = num_WS3
    n_job = num_job
    start_col = input("Masukkan kolom dimulai job 1 WS 1 (cth: B) : ")
    start_row = int(input("Masukkan baris dimulai job 1 WS 1 (cth: 3) : "))

    start_WS1 = [start_col, start_row]
    start_WS2 = [chr(ord(start_WS1[0]) + n_WS1), 3]
    start_WS3 = [chr(ord(start_WS2[0]) + n_WS2), 3]
    start_routing = [chr(ord(start_WS3[0]) + n_WS3), 3]

    # Check all start cells
    # print(start_WS1)
    # print(start_WS2)
    # print(start_WS3)
    # print(start_routing)

    job_WS1 = []
    job_WS2 = []
    job_WS3 = []

    def get_jobs(start_WS, n_job):
        all_num = []
        all_cell = []
        for i in range(n_job):
            all_num.append(str(start_WS[1] + i))
            all_cell.append([start_WS[0], all_num[i]])
        all_string = ["".join(x) for x in all_cell]
        return all_string

    # Get all job data value from worksheet with all job cells
    job_WS1 = [round(ws[x].value, 2)
               for x in get_jobs(start_WS1, n_job)]
    job_WS2 = [round(ws[x].value, 2)
               for x in get_jobs(start_WS2, n_job)]
    job_WS3 = [round(ws[x].value, 2)
               for x in get_jobs(start_WS3, n_job)]
    job_routing = [[int(i) for i in (str(ws[x].value)).split(',')]
                   for x in get_jobs(start_routing, n_job)]

    print("Jobs and routing")
    print("job WS1: ", job_WS1)
    print("job WS2: ", job_WS2)
    print("job WS3: ", job_WS3)
    print("job routing: ", job_routing)

    return job_WS1, job_WS2, job_WS3, job_routing
